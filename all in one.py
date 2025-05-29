
#LOGIN CODE

import subprocess
import cv2
from tkinter import* 
from tkinter import ttk
from PIL import Image,ImageTk
from tkinter import messagebox
from register import Register
import mysql.connector
from train import Train
from student import Student
from train import Train
from attendance import Attendance
from developer import Developer
from helpsupport import Helpsupport
import os
#This are the libraries and other files require to run the code properly.


class Login:
    def __init__(self,root):
        self.root=root
        self.root.title("Login")
        self.root.geometry("1366x768+0+0")   #Here we have declared the title with size of the Frame

        
        # variables 
        self.var_ssq=StringVar()
        self.var_sa=StringVar()
        self.var_pwd=StringVar()

        self.bg=ImageTk.PhotoImage(file=r"d:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/mit.jpg")

        lb1_bg=Label(self.root,image=self.bg)
        lb1_bg.place(x=0,y=0, relwidth=1,relheight=1)

        frame1= Frame(self.root,bg="#002B53")
        frame1.place(x=560,y=170,width=340,height=450)

        img1=Image.open(r"D:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/log1.png")
        img1=img1.resize((100,100),Image.ANTIALIAS)
        self.photoimage1=ImageTk.PhotoImage(img1)
        lb1img1 = Label(image=self.photoimage1,bg="#002B53")
        lb1img1.place(x=690,y=175, width=100,height=100)

        get_str = Label(frame1,text="Teachers Login",font=("times new roman",20,"bold"),fg="White",bg="#002B53")
        get_str.place(x=80,y=100)

        #label1 
        username =lb1= Label(frame1,text="Username:",font=("times new roman",15,"bold"),fg="White",bg="#002B53")
        username.place(x=30,y=160)

        #entry1 
        self.txtuser=ttk.Entry(frame1,font=("times new roman",15,"bold"))
        self.txtuser.place(x=33,y=190,width=270)


        #label2 
        pwd =lb1= Label(frame1,text="Password:",font=("times new roman",15,"bold"),fg="White",bg="#002B53")
        pwd.place(x=30,y=230)

        #entry2 
        self.txtpwd=ttk.Entry(frame1,font=("times new roman",15,"bold"))
        self.txtpwd.place(x=33,y=260,width=270)


        # Creating Button Login
        loginbtn=Button(frame1,command=self.login,text="Login",font=("times new roman",15,"bold"),bd=0,relief=RIDGE,fg="#002B53",bg="white",activeforeground="white",activebackground="#007ACC")
        loginbtn.place(x=33,y=320,width=270,height=35)


        # Creating Button Registration
        loginbtn=Button(frame1,command=self.reg,text="Register",font=("times new roman",10,"bold"),bd=0,relief=RIDGE,fg="white",bg="#002B53",activeforeground="orange",activebackground="#002B53")
        loginbtn.place(x=33,y=370,width=50,height=20)


        # Creating Button Forget
        loginbtn=Button(frame1,command=self.forget_pwd,text="Forget",font=("times new roman",10,"bold"),bd=0,relief=RIDGE,fg="white",bg="#002B53",activeforeground="orange",activebackground="#002B53")
        loginbtn.place(x=90,y=370,width=50,height=20)
        
        '''Now in the above section we have added the paths for the background images, the buttons and there name and also we have created the username and password box
        where the user can enter their mail and password and a new user can register to create there valid credentials.  '''

    # This function is for register window. If the user dont have a valid account they can create it using register button.
    def reg(self):
        self.new_window=Toplevel(self.root)
        self.app=Register(self.new_window)


    def login(self):
        if (self.txtuser.get()=="" or self.txtpwd.get()==""):
            messagebox.showerror("Error","All Field Required!")
        elif(self.txtuser.get()=="admin" and self.txtpwd.get()=="admin"):
            messagebox.showinfo("Sussessfully","Welcome to Attendance Managment System Using Facial Recognition")
        else:
            # messagebox shows an error if the credentials are wrong("Error","Please Check Username or Password !")
            
            conn = mysql.connector.connect(username='root', password='host',host='localhost',database='face_recognition',port=3306)
            mycursor = conn.cursor()
            mycursor.execute("select * from regteach where email=%s and pwd=%s",(
                self.txtuser.get(),
                self.txtpwd.get()
            ))
            row=mycursor.fetchone()
            if row==None:
                messagebox.showerror("Error","Invalid Username and Password!")
            else:
                open_min=messagebox.askyesno("YesNo","Access only Admin")
                if open_min>0:
                    self.new_window=Toplevel(self.root)
                    self.app=Face_Recognition_System(self.new_window)
                else:
                    if not open_min:
                        return
            conn.commit()
            conn.close()
        
        
    #Reset Passowrd Function
    def reset_pass(self):
        if self.var_ssq.get()=="Select":
            messagebox.showerror("Error","Select the Security Question!",parent=self.root2)
        elif(self.var_sa.get()==""):
            messagebox.showerror("Error","Please Enter the Answer!",parent=self.root2)
        elif(self.var_pwd.get()==""):
            messagebox.showerror("Error","Please Enter the New Password!",parent=self.root2)
        else:
            conn = mysql.connector.connect(username='root', password='host',host='localhost',database='face_recognition',port=3306)
            mycursor = conn.cursor()
            query=("select * from regteach where email=%s and ss_que=%s and s_ans=%s")
            value=(self.txtuser.get(),self.var_ssq.get(),self.var_sa.get())
            mycursor.execute(query,value)
            row=mycursor.fetchone()
            if row==None:
                messagebox.showerror("Error","Please Enter the Correct Answer!",parent=self.root2)
            else:
                query=("update regteach set pwd=%s where email=%s")
                value=(self.var_pwd.get(),self.txtuser.get())
                mycursor.execute(query,value)

                conn.commit()
                conn.close()
                messagebox.showinfo("Info","Your password has been reset, Please login with new Password!",parent=self.root2)
                


    #Forget window: In this window the user can reset there password
    def forget_pwd(self):
        if self.txtuser.get()=="":
            messagebox.showerror("Error","Please Enter a Email ID to reset Password!")
        else:
            conn = mysql.connector.connect(username='root', password='host',host='localhost',database='face_recognition',port=3306)
            mycursor = conn.cursor()
            query=("select * from regteach where email=%s")
            value=(self.txtuser.get(),)
            mycursor.execute(query,value)
            row=mycursor.fetchone()
            # print(row)

            if row==None:
                messagebox.showerror("Error","Please Enter the Valid Email ID!")
            else:
                conn.close()
                self.root2=Toplevel()
                self.root2.title("Forget Password")
                self.root2.geometry("400x400+610+170")
                l=Label(self.root2,text="Forget Password",font=("times new roman",30,"bold"),fg="#002B53",bg="#fff")
                l.place(x=0,y=10,relwidth=1)
                
                #Security Question Label
                ssq =lb1= Label(self.root2,text="Select Security Question:",font=("times new roman",15,"bold"),fg="#002B53",bg="#F2F2F2")
                ssq.place(x=70,y=80)

                #Options
                self.combo_security = ttk.Combobox(self.root2,textvariable=self.var_ssq,font=("times new roman",15,"bold"),state="readonly")
                self.combo_security["values"]=("Select","Your Date of Birth","Your Nick Name","Your Favorite Book")
                self.combo_security.current(0)
                self.combo_security.place(x=70,y=110,width=270)


                #Security Question answer label
                sa =lb1= Label(self.root2,text="Security Answer:",font=("times new roman",15,"bold"),fg="#002B53",bg="#F2F2F2")
                sa.place(x=70,y=150)

                #entry oE security Answer
                self.txtpwd=ttk.Entry(self.root2,textvariable=self.var_sa,font=("times new roman",15,"bold"))
                self.txtpwd.place(x=70,y=180,width=270)

                # New Password Label
                new_pwd =lb1= Label(self.root2,text="New Password:",font=("times new roman",15,"bold"),fg="#002B53",bg="#F2F2F2")
                new_pwd.place(x=70,y=220)

                # New password entry 
                self.new_pwd=ttk.Entry(self.root2,textvariable=self.var_pwd,font=("times new roman",15,"bold"))
                self.new_pwd.place(x=70,y=250,width=270)

                # Creating Button of New Password
                loginbtn=Button(self.root2,command=self.reset_pass,text="Reset Password",font=("times new roman",15,"bold"),bd=0,relief=RIDGE,fg="#fff",bg="#002B53",activeforeground="white",activebackground="#007ACC")
                loginbtn.place(x=70,y=300,width=270,height=35)


            

# Main Program 

class Face_Recognition_System:
    def __init__(self,root):
        self.root=root
        self.root.geometry("1366x768+0+0")
        self.root.title("Face_Recogonition_System")  #The face recognition name is given to the code

        #Setting up the Image labels and adding background image
        
        img=Image.open(r"d:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/mitname.jpg")
        img=img.resize((1366,130),Image.ANTIALIAS)
        self.photoimg=ImageTk.PhotoImage(img)

        # set image as lable
        f_lb1 = Label(self.root,image=self.photoimg)
        f_lb1.place(x=0,y=0,width=1366,height=130)

        # backgorund image 
        bg1=Image.open(r"d:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/back.jpg")
        bg1=bg1.resize((1366,768),Image.ANTIALIAS)
        self.photobg1=ImageTk.PhotoImage(bg1)

        # set image as lable
        bg_img = Label(self.root,image=self.photobg1)
        bg_img.place(x=0,y=130,width=1366,height=768)


        #title section
        title_lb1 = Label(bg_img,text="Attendance Managment System Using Facial Recognition",font=("verdana",30,"bold"),bg="white",fg="navyblue")
        title_lb1.place(x=0,y=0,width=1366,height=45)

        # In the below section the process for button creation is given
    
        # student button 1
        std_img_btn=Image.open(r"D:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/std1.jpg")
        std_img_btn=std_img_btn.resize((180,180),Image.ANTIALIAS)
        self.std_img1=ImageTk.PhotoImage(std_img_btn)

        std_b1 = Button(bg_img,command=self.student_pannels,image=self.std_img1,cursor="hand2")
        std_b1.place(x=250,y=100,width=180,height=180)

        std_b1_1 = Button(bg_img,command=self.student_pannels,text="Student Pannel",cursor="hand2",font=("tahoma",15,"bold"),bg="white",fg="navyblue")
        std_b1_1.place(x=250,y=280,width=180,height=45)

        # Detect Face  button 2
        det_img_btn=Image.open(r"D:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/det1.jpg")
        det_img_btn=det_img_btn.resize((180,180),Image.ANTIALIAS)
        self.det_img1=ImageTk.PhotoImage(det_img_btn)

        det_b1 = Button(bg_img,command=self.face_rec,image=self.det_img1,cursor="hand2",)
        det_b1.place(x=480,y=100,width=180,height=180)

        det_b1_1 = Button(bg_img,command=self.face_rec,text="Face Detector",cursor="hand2",font=("tahoma",15,"bold"),bg="white",fg="navyblue")
        det_b1_1.place(x=480,y=280,width=180,height=45)

         # Attendance System  button 3
        att_img_btn=Image.open(r"D:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/att.jpg")
        att_img_btn=att_img_btn.resize((180,180),Image.ANTIALIAS)
        self.att_img1=ImageTk.PhotoImage(att_img_btn)

        att_b1 = Button(bg_img,command=self.attendance_pannel,image=self.att_img1,cursor="hand2",)
        att_b1.place(x=710,y=100,width=180,height=180)

        att_b1_1 = Button(bg_img,command=self.attendance_pannel,text="Attendance",cursor="hand2",font=("tahoma",15,"bold"),bg="white",fg="navyblue")
        att_b1_1.place(x=710,y=280,width=180,height=45)

         # Help  Support  button 4
        hlp_img_btn=Image.open(r"D:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/hlp.jpg")
        hlp_img_btn=hlp_img_btn.resize((180,180),Image.ANTIALIAS)
        self.hlp_img1=ImageTk.PhotoImage(hlp_img_btn)

        hlp_b1 = Button(bg_img,command=self.helpSupport,image=self.hlp_img1,cursor="hand2",)
        hlp_b1.place(x=940,y=100,width=180,height=180)

        hlp_b1_1 = Button(bg_img,command=self.helpSupport,text="Help Support",cursor="hand2",font=("tahoma",15,"bold"),bg="white",fg="navyblue")
        hlp_b1_1.place(x=940,y=280,width=180,height=45)

    
         # Train   button 5
        tra_img_btn=Image.open(r"D:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/tra1.jpg")
        tra_img_btn=tra_img_btn.resize((180,180),Image.ANTIALIAS)
        self.tra_img1=ImageTk.PhotoImage(tra_img_btn)

        tra_b1 = Button(bg_img,command=self.train_pannels,image=self.tra_img1,cursor="hand2",)
        tra_b1.place(x=250,y=330,width=180,height=180)

        tra_b1_1 = Button(bg_img,command=self.train_pannels,text="Data Train",cursor="hand2",font=("tahoma",15,"bold"),bg="white",fg="navyblue")
        tra_b1_1.place(x=250,y=510,width=180,height=45)

        # Photo   button 6
        pho_img_btn=Image.open(r"d:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/camera.jpg")
        pho_img_btn=pho_img_btn.resize((180,180),Image.ANTIALIAS)
        self.pho_img1=ImageTk.PhotoImage(pho_img_btn)

        pho_b1 = Button(bg_img,command=self.open_img,image=self.pho_img1,cursor="hand2",)
        pho_b1.place(x=480,y=330,width=180,height=180)

        pho_b1_1 = Button(bg_img,command=self.open_img,text="Capture Image",cursor="hand2",font=("tahoma",15,"bold"),bg="white",fg="navyblue")
        pho_b1_1.place(x=480,y=510,width=180,height=45)

        # Developers   button 7
        dev_img_btn=Image.open(r"D:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/dev.jpg")
        dev_img_btn=dev_img_btn.resize((180,180),Image.ANTIALIAS)
        self.dev_img1=ImageTk.PhotoImage(dev_img_btn)

        dev_b1 = Button(bg_img,command=self.developr,image=self.dev_img1,cursor="hand2",)
        dev_b1.place(x=710,y=330,width=180,height=180)

        dev_b1_1 = Button(bg_img,command=self.developr,text="Developers",cursor="hand2",font=("tahoma",15,"bold"),bg="white",fg="navyblue")
        dev_b1_1.place(x=710,y=510,width=180,height=45)

        # exit   button 8
        exi_img_btn=Image.open(r"D:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/exi.jpg")
        exi_img_btn=exi_img_btn.resize((180,180),Image.ANTIALIAS)
        self.exi_img1=ImageTk.PhotoImage(exi_img_btn)

        exi_b1 = Button(bg_img,command=self.Close,image=self.exi_img1,cursor="hand2",)
        exi_b1.place(x=940,y=330,width=180,height=180)

        exi_b1_1 = Button(bg_img,command=self.Close,text="Exit",cursor="hand2",font=("tahoma",15,"bold"),bg="white",fg="navyblue")
        exi_b1_1.place(x=940,y=510,width=180,height=45)

#Funtion for Opening Images Folder
    def open_img(self):
        capture = cv2.VideoCapture(0)  
        
        ret, frame = capture.read()

        if ret:
            save_directory = "D:/studentsdetails/"
            image_name = "test.jpg"         # Here we have added the image path and its name so that it should open this image and recognize the faces in that image
            image_path = os.path.join(save_directory, image_name)

            if os.path.exists(image_path):
                os.remove(image_path)

            cv2.imwrite(image_path, frame)
            print("Image saved:", image_path)
        else:
            print("Failed to capture image.")

        capture.release()  

# Functions Buttons
    def student_pannels(self):
        self.new_window=Toplevel(self.root)
        self.app=Student(self.new_window)

    def train_pannels(self):
        self.new_window=Toplevel(self.root)
        self.app=Train(self.new_window)
    
    def face_rec(self):
        script_directory = r'D:/studentsdetails'

        script_name = 'd:/studentsdetails/faces.py'

        script_path = os.path.join(script_directory, script_name)

        os.chdir(script_directory)

        subprocess.Popen(['python', script_path],shell=True)
    
    def attendance_pannel(self):
        self.new_window=Toplevel(self.root)
        self.app=Attendance(self.new_window)
    
    def developr(self):
        self.new_window=Toplevel(self.root)
        self.app=Developer(self.new_window)
    
    def helpSupport(self):
        self.new_window=Toplevel(self.root)
        self.app=Helpsupport(self.new_window)
        
    def Close(self):
        root.destroy()

    
if __name__ == "__main__":
    root=Tk()
    app=Login(root)
    root.mainloop()
    
#CODE FOR REGISTRATION 

from tkinter import* 
from tkinter import ttk
from PIL import Image,ImageTk
from tkinter import messagebox
import mysql.connector
#This are the libraries and other files require to run the code properly.

class Register:
    def __init__(self,root):
        self.root=root
        self.root.title("Register")
        self.root.geometry("1366x768+0+0")   #The code is named as register and the size is define

        # Here we declared the Variables
        self.var_fname=StringVar()
        self.var_lname=StringVar()
        self.var_cnum=StringVar()
        self.var_email=StringVar()
        self.var_ssq=StringVar()
        self.var_sa=StringVar()
        self.var_pwd=StringVar()
        self.var_cpwd=StringVar()
        self.var_check=IntVar()

#From this part we and the images created box and tile and buttons for registration process
        self.bg=ImageTk.PhotoImage(file=r"D:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/bgReg.jpg")
        
        lb1_bg=Label(self.root,image=self.bg)
        lb1_bg.place(x=0,y=0, relwidth=1,relheight=1)

        frame= Frame(self.root,bg="#F2F2F2")
        frame.place(x=100,y=80,width=900,height=580)
        

        img1=Image.open(r"D:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/reg1.png")
        img1=img1.resize((450,100),Image.ANTIALIAS)
        self.photoimage1=ImageTk.PhotoImage(img1)
        lb1img1 = Label(image=self.photoimage1,bg="#F2F2F2")
        lb1img1.place(x=300,y=100, width=500,height=100)
        

        get_str = Label(frame,text="Registration",font=("times new roman",30,"bold"),fg="#002B53",bg="#F2F2F2")
        get_str.place(x=350,y=130)

        #Label for First name
        fname =lb1= Label(frame,text="First Name:",font=("times new roman",15,"bold"),fg="#002B53",bg="#F2F2F2")
        fname.place(x=100,y=200)

        #Entry of first name
        self.txtuser=ttk.Entry(frame,textvariable=self.var_fname,font=("times new roman",15,"bold"))
        self.txtuser.place(x=103,y=225,width=270)


        #Label for last name
        lname =lb1= Label(frame,text="Last Name:",font=("times new roman",15,"bold"),fg="#002B53",bg="#F2F2F2")
        lname.place(x=100,y=270)

        #Entry of last name
        self.txtpwd=ttk.Entry(frame,textvariable=self.var_lname,font=("times new roman",15,"bold"))
        self.txtpwd.place(x=103,y=295,width=270)

       

        #Label for contact number
        cnum =lb1= Label(frame,text="Contact No:",font=("times new roman",15,"bold"),fg="#002B53",bg="#F2F2F2")
        cnum.place(x=530,y=200)

        #Entry for contact number 
        self.txtuser=ttk.Entry(frame,textvariable=self.var_cnum,font=("times new roman",15,"bold"))
        self.txtuser.place(x=533,y=225,width=270)


        #Label for email
        email =lb1= Label(frame,text="Email:",font=("times new roman",15,"bold"),fg="#002B53",bg="#F2F2F2")
        email.place(x=530,y=270)

        #Entry for email
        self.txtpwd=ttk.Entry(frame,textvariable=self.var_email,font=("times new roman",15,"bold"))
        self.txtpwd.place(x=533,y=295,width=270)

       

        #Label for security question  
        ssq =lb1= Label(frame,text="Select Security Question:",font=("times new roman",15,"bold"),fg="#002B53",bg="#F2F2F2")
        ssq.place(x=100,y=350)

        #Options for security question
        self.combo_security = ttk.Combobox(frame,textvariable=self.var_ssq,font=("times new roman",15,"bold"),state="readonly")
        self.combo_security["values"]=("Select","Your Date of Birth","Your Nick Name","Your Favorite Book")
        self.combo_security.current(0)
        self.combo_security.place(x=103,y=375,width=270)


        #Label of security question answer
        sa =lb1= Label(frame,text="Security Answer:",font=("times new roman",15,"bold"),fg="#002B53",bg="#F2F2F2")
        sa.place(x=100,y=420)

        #Entry of security question answer
        self.txtpwd=ttk.Entry(frame,textvariable=self.var_sa,font=("times new roman",15,"bold"))
        self.txtpwd.place(x=103,y=445,width=270)


        #Label for password
        pwd =lb1= Label(frame,text="Password:",font=("times new roman",15,"bold"),fg="#002B53",bg="#F2F2F2")
        pwd.place(x=530,y=350)

        #Entry of password
        self.txtuser=ttk.Entry(frame,textvariable=self.var_pwd,font=("times new roman",15,"bold"))
        self.txtuser.place(x=533,y=375,width=270)


        #Label for confirm password
        cpwd =lb1= Label(frame,text="Confirm Password:",font=("times new roman",15,"bold"),fg="#002B53",bg="#F2F2F2")
        cpwd.place(x=530,y=420)

        #Entry of confirm password
        self.txtpwd=ttk.Entry(frame,textvariable=self.var_cpwd,font=("times new roman",15,"bold"))
        self.txtpwd.place(x=533,y=445,width=270)

        # Checkbutton
        checkbtn = Checkbutton(frame,variable=self.var_check,text="I Agree the Terms & Conditions",font=("times new roman",13,"bold"),fg="#002B53",bg="#F2F2F2")
        checkbtn.place(x=100,y=480,width=270)


        # Creating Button for Register
        loginbtn=Button(frame,command=self.reg,text="Register",font=("times new roman",15,"bold"),bd=0,relief=RIDGE,fg="#fff",bg="#002B53",activeforeground="white",activebackground="#007ACC")
        loginbtn.place(x=103,y=510,width=270,height=35)

        # Creating Button for Login
        loginbtn=Button(frame,text="Login",font=("times new roman",15,"bold"),bd=0,relief=RIDGE,fg="#fff",bg="#002B53",activeforeground="white",activebackground="#007ACC")
        loginbtn.place(x=533,y=510,width=270,height=35)



#This section is for checking everything and if anything is missing then it display the given message

    def reg(self):
        if (self.var_fname.get()=="" or self.var_lname.get()=="" or self.var_cnum.get()=="" or self.var_email.get()=="" or self.var_ssq.get()=="Select" or self.var_sa.get()=="" or self.var_pwd.get()=="" or self.var_cpwd.get()==""):
            messagebox.showerror("Error","All Field Required!")
        elif(self.var_pwd.get() != self.var_cpwd.get()):
            messagebox.showerror("Error","Please Enter Password & Confirm Password are Same!")
        elif(self.var_check.get()==0):
            messagebox.showerror("Error","Please Check the Agree Terms and Conditons!")
        else:
        
            try:
                conn = mysql.connector.connect(username='root', password='host',host='localhost',database='face_recognition',port=3306)
                mycursor = conn.cursor()
                query=("select * from regteach where email=%s")
                value=(self.var_email.get(),)
                mycursor.execute(query,value)
                row=mycursor.fetchone()
                if row!=None:
                    messagebox.showerror("Error","User already exist,please try another email")
                else:
                    mycursor.execute("insert into regteach values(%s,%s,%s,%s,%s,%s,%s)",(
                    self.var_fname.get(),
                    self.var_lname.get(),
                    self.var_cnum.get(),
                    self.var_email.get(),
                    self.var_ssq.get(),
                    self.var_sa.get(),
                    self.var_pwd.get()
                    ))

                    conn.commit()
                    conn.close()
                    messagebox.showinfo("Success","Successfully Registerd!",parent=self.root) #Once successfull registration is done it will display the given message
            except Exception as es:
                messagebox.showerror("Error",f"Due to: {str(es)}",parent=self.root)


if __name__ == "__main__":
    root=Tk()
    app=Register(root)
    root.mainloop()
    

#MAIN CODE

import subprocess
import cv2
from tkinter import*
from tkinter import ttk
from train import Train
from PIL import Image,ImageTk
from student import Student
from train import Train
from face_recognition import Face_Recognition
from attendance import Attendance
from developer import Developer
import os
from helpsupport import Helpsupport
#This are the libraries and other files require to run the code properly.


class Face_Recognition_System:
    def __init__(self,root):
        self.root=root
        self.root.geometry("1366x768+0+0")
        self.root.title("Face_Recogonition_System")    #The face recognition name is given to the code

        #Setting up the Image labels, Buttons and adding background image 
        
        #Header image
        img=Image.open(r"d:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/mitname.jpg")
        img=img.resize((1366,130),Image.ANTIALIAS)
        self.photoimg=ImageTk.PhotoImage(img)

        # set image as lable
        f_lb1 = Label(self.root,image=self.photoimg)
        f_lb1.place(x=0,y=0,width=1366,height=130)

        # backgorund image 
        bg1=Image.open(r"d:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/back.jpg")
        bg1=bg1.resize((1366,768),Image.ANTIALIAS)
        self.photobg1=ImageTk.PhotoImage(bg1)

        # set image as lable
        bg_img = Label(self.root,image=self.photobg1)
        bg_img.place(x=0,y=130,width=1366,height=768)


        #title section
        title_lb1 = Label(bg_img,text="Automated Attendance System",font=("verdana",27,"bold"),bg="white",fg="navyblue")
        title_lb1.place(x=0,y=0,width=1366,height=45)

# In the below section the process for button creation is given

        # student button 1
        std_img_btn=Image.open(r"D:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/std1.jpg")
        std_img_btn=std_img_btn.resize((180,180),Image.ANTIALIAS)
        self.std_img1=ImageTk.PhotoImage(std_img_btn)

        std_b1 = Button(bg_img,command=self.student_pannels,image=self.std_img1,cursor="hand2")
        std_b1.place(x=250,y=100,width=180,height=180)

        std_b1_1 = Button(bg_img,command=self.student_pannels,text="Student Pannel",cursor="hand2",font=("tahoma",15,"bold"),bg="white",fg="navyblue")
        std_b1_1.place(x=250,y=280,width=180,height=45)

        # Detect Face  button 2
        det_img_btn=Image.open(r"D:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/det1.jpg")
        det_img_btn=det_img_btn.resize((180,180),Image.ANTIALIAS)
        self.det_img1=ImageTk.PhotoImage(det_img_btn)

        det_b1 = Button(bg_img,command=self.face_rec,image=self.det_img1,cursor="hand2",)
        det_b1.place(x=480,y=100,width=180,height=180)

        det_b1_1 = Button(bg_img,command=self.face_rec,text="Face Detector",cursor="hand2",font=("tahoma",15,"bold"),bg="white",fg="navyblue")
        det_b1_1.place(x=480,y=280,width=180,height=45)

         # Attendance System  button 3
        att_img_btn=Image.open(r"D:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/att.jpg")
        att_img_btn=att_img_btn.resize((180,180),Image.ANTIALIAS)
        self.att_img1=ImageTk.PhotoImage(att_img_btn)

        att_b1 = Button(bg_img,command=self.attendance_pannel,image=self.att_img1,cursor="hand2",)
        att_b1.place(x=710,y=100,width=180,height=180)

        att_b1_1 = Button(bg_img,command=self.attendance_pannel,text="Attendance",cursor="hand2",font=("tahoma",15,"bold"),bg="white",fg="navyblue")
        att_b1_1.place(x=710,y=280,width=180,height=45)

         # Help  Support  button 4
        hlp_img_btn=Image.open(r"D:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/hlp.jpg")
        hlp_img_btn=hlp_img_btn.resize((180,180),Image.ANTIALIAS)
        self.hlp_img1=ImageTk.PhotoImage(hlp_img_btn)

        hlp_b1 = Button(bg_img,command=self.helpSupport,image=self.hlp_img1,cursor="hand2",)
        hlp_b1.place(x=940,y=100,width=180,height=180)

        hlp_b1_1 = Button(bg_img,command=self.helpSupport,text="Help Support",cursor="hand2",font=("tahoma",15,"bold"),bg="white",fg="navyblue")
        hlp_b1_1.place(x=940,y=280,width=180,height=45)

      
         # Train   button 5
        tra_img_btn=Image.open(r"D:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/tra1.jpg")
        tra_img_btn=tra_img_btn.resize((180,180),Image.ANTIALIAS)
        self.tra_img1=ImageTk.PhotoImage(tra_img_btn)

        tra_b1 = Button(bg_img,command=self.train_pannels,image=self.tra_img1,cursor="hand2",)
        tra_b1.place(x=250,y=330,width=180,height=180)

        tra_b1_1 = Button(bg_img,command=self.train_pannels,text="Data Train",cursor="hand2",font=("tahoma",15,"bold"),bg="white",fg="navyblue")
        tra_b1_1.place(x=250,y=510,width=180,height=45)

        # Photo   button 6
        pho_img_btn=Image.open(r"d:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/camera.jpg")
        pho_img_btn=pho_img_btn.resize((180,180),Image.ANTIALIAS)
        self.pho_img1=ImageTk.PhotoImage(pho_img_btn)

        pho_b1 = Button(bg_img,command=self.open_img,image=self.pho_img1,cursor="hand2",)
        pho_b1.place(x=480,y=330,width=180,height=180)

        pho_b1_1 = Button(bg_img,command=self.open_img,text="Capture Image",cursor="hand2",font=("tahoma",15,"bold"),bg="white",fg="navyblue")
        pho_b1_1.place(x=480,y=510,width=180,height=45)

        # Developers   button 7
        dev_img_btn=Image.open(r"D:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/dev.jpg")
        dev_img_btn=dev_img_btn.resize((180,180),Image.ANTIALIAS)
        self.dev_img1=ImageTk.PhotoImage(dev_img_btn)

        dev_b1 = Button(bg_img,command=self.developr,image=self.dev_img1,cursor="hand2",)
        dev_b1.place(x=710,y=330,width=180,height=180)

        dev_b1_1 = Button(bg_img,command=self.developr,text="Developers",cursor="hand2",font=("tahoma",15,"bold"),bg="white",fg="navyblue")
        dev_b1_1.place(x=710,y=510,width=180,height=45)

        # exit   button 8
        exi_img_btn=Image.open(r"D:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/exi.jpg")
        exi_img_btn=exi_img_btn.resize((180,180),Image.ANTIALIAS)
        self.exi_img1=ImageTk.PhotoImage(exi_img_btn)

        exi_b1 = Button(bg_img,command=self.Close,image=self.exi_img1,cursor="hand2",)
        exi_b1.place(x=940,y=330,width=180,height=180)

        exi_b1_1 = Button(bg_img,command=self.Close,text="Exit",cursor="hand2",font=("tahoma",15,"bold"),bg="white",fg="navyblue")
        exi_b1_1.place(x=940,y=510,width=180,height=45)
        

# Funtion for Opening Images Folder
    def open_img(self):
        capture = cv2.VideoCapture(0)  

        ret, frame = capture.read()

        if ret:
            save_directory = "D:/studentsdetails/"
            image_name = "test.jpg"     # Here we have added the image path and its name so that it should open this image and recognize the faces in that image
            image_path = os.path.join(save_directory, image_name)

            if os.path.exists(image_path):
                os.remove(image_path)

            # Save the captured frame as an image
            cv2.imwrite(image_path, frame)
            print("Image saved:", image_path)
        else:
            print("Failed to capture image.")

        capture.release() 

#Functions Buttons
    def student_pannels(self):
        self.new_window=Toplevel(self.root)
        self.app=Student(self.new_window)

    def train_pannels(self):
        self.new_window=Toplevel(self.root)
        self.app=Train(self.new_window)
    
    def face_rec(self):
       script_directory = r'D:/studentsdetails'

       script_name = 'd:/studentsdetails/faces.py'

       script_path = os.path.join(script_directory, script_name)

       os.chdir(script_directory)

       subprocess.Popen(['python', script_path],shell=True)
       
    def attendance_pannel(self):
        self.new_window=Toplevel(self.root)
        self.app=Attendance(self.new_window)
    
    def developr(self):
        self.new_window=Toplevel(self.root)
        self.app=Developer(self.new_window)
    
    def helpSupport(self):
        self.new_window=Toplevel(self.root)
        self.app=Helpsupport(self.new_window)

    def Close(self):
        root.destroy()

if __name__ == "__main__":
    root=Tk()
    obj=Face_Recognition_System(root)
    root.mainloop()


#STUDENT PANEL CODE
#The code is for filling all the details of student

from tkinter import* 
from tkinter import ttk
from PIL import Image,ImageTk
from tkinter import messagebox
import mysql.connector
import cv2
#This are the libraries and other files require to run the code properly.


class Student:
    def __init__(self,root):
        self.root=root
        self.root.geometry("1366x768+0+0")
        self.root.title("Student Pannel") 

        #Here we declared the variables
        self.var_dep=StringVar()
        self.var_course=StringVar()
        self.var_year=StringVar()
        self.var_semester=StringVar()
        self.var_std_id=StringVar()
        self.var_std_name=StringVar()
        self.var_div=StringVar()
        self.var_roll=StringVar()
        self.var_gender=StringVar()
        self.var_dob=StringVar()
        self.var_email=StringVar()
        self.var_mob=StringVar()
        self.var_address=StringVar()
        self.var_teacher=StringVar()

#In this section we are addinng the header,background,boxes and labels 
       
        # Header image  
        img=Image.open(r"d:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/mitname.jpg")
        img=img.resize((1366,130),Image.ANTIALIAS)
        self.photoimg=ImageTk.PhotoImage(img)  #In this section we added the header image

        # Set the above image as a label
        f_lb1 = Label(self.root,image=self.photoimg)
        f_lb1.place(x=0,y=0,width=1366,height=130)

         # Backgorund image 
        bg1=Image.open(r"D:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/bg3.jpg")
        bg1=bg1.resize((1366,768),Image.ANTIALIAS)
        self.photobg1=ImageTk.PhotoImage(bg1)

        # Set the above image as a label
        bg_img = Label(self.root,image=self.photobg1)
        bg_img.place(x=0,y=130,width=1366,height=768)


        # Title section
        title_lb1 = Label(bg_img,text="Welcome to Student Pannel",font=("verdana",30,"bold"),bg="white",fg="navyblue")
        title_lb1.place(x=0,y=0,width=1366,height=45)  #Added the title

        # Creating Frame 
        main_frame = Frame(bg_img,bd=2,bg="white") 
        main_frame.place(x=5,y=55,width=1355,height=510) #Here bd means border

        # Left Label Frame 
        left_frame = LabelFrame(main_frame,bd=2,bg="white",relief=RIDGE,text="Student Details",font=("verdana",12,"bold"),fg="navyblue")
        left_frame.place(x=10,y=10,width=660,height=480)

        # Current Course 
        current_course_frame = LabelFrame(left_frame,bd=2,bg="white",relief=RIDGE,text="Current Course",font=("verdana",12,"bold"),fg="navyblue")
        current_course_frame.place(x=10,y=5,width=635,height=150)

        # Label for Department
        dep_label=Label(current_course_frame,text="Department",font=("verdana",12,"bold"),bg="white",fg="navyblue")
        dep_label.grid(row=0,column=0,padx=5,pady=15)

        #Added options for Department
        dep_combo=ttk.Combobox(current_course_frame,textvariable=self.var_dep,width=15,font=("verdana",12,"bold"),state="readonly")
        dep_combo["values"]=("Select Department","Electronics and Telecommunication","Mechanical","Computer","Civil","Electrical")
        dep_combo.current(0)
        dep_combo.grid(row=0,column=1,padx=5,pady=15,sticky=W)


        #Label for Course
        cou_label=Label(current_course_frame,text="Course",font=("verdana",12,"bold"),bg="white",fg="navyblue")
        cou_label.grid(row=0,column=2,padx=5,pady=15)

        #Added options for the course 
        cou_combo=ttk.Combobox(current_course_frame,textvariable=self.var_course,width=15,font=("verdana",12,"bold"),state="readonly")
        cou_combo["values"]=("Select Course","VLSI","Pyrhon","JAVA","Siganl and System","Maths")
        cou_combo.current(0)
        cou_combo.grid(row=0,column=3,padx=5,pady=15,sticky=W)
    

        #Label for Year
        year_label=Label(current_course_frame,text="Year",font=("verdana",12,"bold"),bg="white",fg="navyblue")
        year_label.grid(row=1,column=0,padx=5,sticky=W)

        #Options for Year in which the student is studying 
        year_combo=ttk.Combobox(current_course_frame,textvariable=self.var_year,width=15,font=("verdana",12,"bold"),state="readonly")
        year_combo["values"]=("Select Year","2017-21","2018-22","2019-23","2020-24","2021-25")
        year_combo.current(0)
        year_combo.grid(row=1,column=1,padx=5,pady=15,sticky=W)


        #Label for Semester 
        year_label=Label(current_course_frame,text="Semester",font=("verdana",12,"bold"),bg="white",fg="navyblue")
        year_label.grid(row=1,column=2,padx=5,sticky=W)

        #Different semester options
        year_combo=ttk.Combobox(current_course_frame,textvariable=self.var_semester,width=15,font=("verdana",12,"bold"),state="readonly")
        year_combo["values"]=("Select Semester","Semester-1","Semester-2","Semester-3","Semester-4","Semester-5","Semester-6","Semester-7","Semester-8")
        year_combo.current(0)
        year_combo.grid(row=1,column=3,padx=5,pady=15,sticky=W)

        #Class Student Information
        class_Student_frame = LabelFrame(left_frame,bd=2,bg="white",relief=RIDGE,text="Class Student Information",font=("verdana",12,"bold"),fg="navyblue")
        class_Student_frame.place(x=10,y=160,width=635,height=230)

        #Student Name
        student_name_label = Label(class_Student_frame,text="Std-Name:",font=("verdana",12,"bold"),fg="navyblue",bg="white")
        student_name_label.grid(row=0,column=0,padx=5,pady=5,sticky=W)

        student_name_entry = ttk.Entry(class_Student_frame,textvariable=self.var_std_id,width=15,font=("verdana",12,"bold"))
        student_name_entry.grid(row=0,column=1,padx=5,pady=5,sticky=W)

        #Student roll-no
        student_roll_label = Label(class_Student_frame,text="Roll-NO:",font=("verdana",12,"bold"),fg="navyblue",bg="white")
        student_roll_label.grid(row=0,column=2,padx=5,pady=5,sticky=W)

        student_roll_entry = ttk.Entry(class_Student_frame,textvariable=self.var_std_name,width=15,font=("verdana",12,"bold"))
        student_roll_entry.grid(row=0,column=3,padx=5,pady=5,sticky=W)

        #Class Didvision
        student_div_label = Label(class_Student_frame,text="Class Division:",font=("verdana",12,"bold"),fg="navyblue",bg="white")
        student_div_label.grid(row=1,column=0,padx=5,pady=5,sticky=W)

        div_combo=ttk.Combobox(class_Student_frame,textvariable=self.var_div,width=13,font=("verdana",12,"bold"),state="readonly")
        div_combo["values"]=("Morning","Evening")
        div_combo.current(0)
        div_combo.grid(row=1,column=1,padx=5,pady=5,sticky=W)

        #Student ID
        studentId_label = Label(class_Student_frame,text="Std-ID:",font=("verdana",12,"bold"),fg="navyblue",bg="white")
        studentId_label.grid(row=1,column=2,padx=5,pady=5,sticky=W)

        studentId_entry = ttk.Entry(class_Student_frame,textvariable=self.var_roll,width=15,font=("verdana",12,"bold"))
        studentId_entry.grid(row=1,column=3,padx=5,pady=5,sticky=W)

        #Gender
        student_gender_label = Label(class_Student_frame,text="Gender:",font=("verdana",12,"bold"),fg="navyblue",bg="white")
        student_gender_label.grid(row=2,column=0,padx=5,pady=5,sticky=W)

        #Options for selsction of Gender
        gender_combo=ttk.Combobox(class_Student_frame,textvariable=self.var_gender,width=13,font=("verdana",12,"bold"),state="readonly")
        gender_combo["values"]=("Male","Female","Others")
        gender_combo.current(0)
        gender_combo.grid(row=2,column=1,padx=5,pady=5,sticky=W)

        #Date of Birth
        student_dob_label = Label(class_Student_frame,text="DOB:",font=("verdana",12,"bold"),fg="navyblue",bg="white")
        student_dob_label.grid(row=2,column=2,padx=5,pady=5,sticky=W)

        student_dob_entry = ttk.Entry(class_Student_frame,textvariable=self.var_dob,width=15,font=("verdana",12,"bold"))
        student_dob_entry.grid(row=2,column=3,padx=5,pady=5,sticky=W)

        #Email
        student_email_label = Label(class_Student_frame,text="Email:",font=("verdana",12,"bold"),fg="navyblue",bg="white")
        student_email_label.grid(row=3,column=0,padx=5,pady=5,sticky=W)

        student_email_entry = ttk.Entry(class_Student_frame,textvariable=self.var_email,width=15,font=("verdana",12,"bold"))
        student_email_entry.grid(row=3,column=1,padx=5,pady=5,sticky=W)

        #Phone Number
        student_mob_label = Label(class_Student_frame,text="Mob-No:",font=("verdana",12,"bold"),fg="navyblue",bg="white")
        student_mob_label.grid(row=3,column=2,padx=5,pady=5,sticky=W)

        student_mob_entry = ttk.Entry(class_Student_frame,textvariable=self.var_mob,width=15,font=("verdana",12,"bold"))
        student_mob_entry.grid(row=3,column=3,padx=5,pady=5,sticky=W)

        #Address
        student_address_label = Label(class_Student_frame,text="Address:",font=("verdana",12,"bold"),fg="navyblue",bg="white")
        student_address_label.grid(row=4,column=0,padx=5,pady=5,sticky=W)

        student_address_entry = ttk.Entry(class_Student_frame,textvariable=self.var_address,width=15,font=("verdana",12,"bold"))
        student_address_entry.grid(row=4,column=1,padx=5,pady=5,sticky=W)

        #Teacher Name
        student_tutor_label = Label(class_Student_frame,text="Tutor Name:",font=("verdana",12,"bold"),fg="navyblue",bg="white")
        student_tutor_label.grid(row=4,column=2,padx=5,pady=5,sticky=W)

        student_tutor_entry = ttk.Entry(class_Student_frame,textvariable=self.var_teacher,width=15,font=("verdana",12,"bold"))
        student_tutor_entry.grid(row=4,column=3,padx=5,pady=5,sticky=W)

        #Take Photo sample button
        self.var_radio1=StringVar()
        radiobtn1=ttk.Radiobutton(class_Student_frame,text="Take Photo Sample",variable=self.var_radio1,value="Yes")
        radiobtn1.grid(row=5,column=0,padx=5,pady=5,sticky=W)

        radiobtn1=ttk.Radiobutton(class_Student_frame,text="No Photo Sample",variable=self.var_radio1,value="No")
        radiobtn1.grid(row=5,column=1,padx=5,pady=5,sticky=W)

        #Button Frame for all the button
        btn_frame = Frame(left_frame,bd=2,bg="white",relief=RIDGE)
        btn_frame.place(x=10,y=390,width=635,height=60)

        #Save button
        save_btn=Button(btn_frame,command=self.add_data,text="Save",width=7,font=("verdana",12,"bold"),fg="white",bg="navyblue")
        save_btn.grid(row=0,column=0,padx=5,pady=10,sticky=W)

        #Update button
        update_btn=Button(btn_frame,command=self.update_data,text="Update",width=7,font=("verdana",12,"bold"),fg="white",bg="navyblue")
        update_btn.grid(row=0,column=1,padx=5,pady=8,sticky=W)

        #Delete button
        del_btn=Button(btn_frame,command=self.delete_data,text="Delete",width=7,font=("verdana",12,"bold"),fg="white",bg="navyblue")
        del_btn.grid(row=0,column=2,padx=5,pady=10,sticky=W)

        #Reset button
        reset_btn=Button(btn_frame,command=self.reset_data,text="Reset",width=7,font=("verdana",12,"bold"),fg="white",bg="navyblue")
        reset_btn.grid(row=0,column=3,padx=5,pady=10,sticky=W)

        #Take photo button
        take_photo_btn=Button(btn_frame,command=self.generate_dataset,text="Take Pic",width=9,font=("verdana",12,"bold"),fg="white",bg="navyblue")
        take_photo_btn.grid(row=0,column=4,padx=5,pady=10,sticky=W)

        #Update photo button
        update_photo_btn=Button(btn_frame,text="Update Pic",width=9,font=("verdana",12,"bold"),fg="white",bg="navyblue")
        update_photo_btn.grid(row=0,column=5,padx=5,pady=10,sticky=W)


#Now the above section we created on left side and now we added some more options on right side

        # Right Label Frame 
        right_frame = LabelFrame(main_frame,bd=2,bg="white",relief=RIDGE,text="Student Details",font=("verdana",12,"bold"),fg="navyblue")
        right_frame.place(x=680,y=10,width=660,height=480)

        #Searching System in Right Label Frame 
        search_frame = LabelFrame(right_frame,bd=2,bg="white",relief=RIDGE,text="Search System",font=("verdana",12,"bold"),fg="navyblue")
        search_frame.place(x=10,y=5,width=635,height=80)

        #Phone Number
        search_label = Label(search_frame,text="Search:",font=("verdana",12,"bold"),fg="navyblue",bg="white")
        search_label.grid(row=0,column=0,padx=5,pady=5,sticky=W)
        self.var_searchTX=StringVar()
 
        #Option
        search_combo=ttk.Combobox(search_frame,textvariable=self.var_searchTX,width=12,font=("verdana",12,"bold"),state="readonly")
        search_combo["values"]=("Select","Roll-No")
        search_combo.current(0)
        search_combo.grid(row=0,column=1,padx=5,pady=15,sticky=W) #Using the roll-no this section can search the student profile

        self.var_search=StringVar()
        search_entry = ttk.Entry(search_frame,textvariable=self.var_search,width=12,font=("verdana",12,"bold"))
        search_entry.grid(row=0,column=2,padx=5,pady=5,sticky=W)

        search_btn=Button(search_frame,command=self.search_data,text="Search",width=9,font=("verdana",12,"bold"),fg="white",bg="navyblue")
        search_btn.grid(row=0,column=3,padx=5,pady=10,sticky=W)

        showAll_btn=Button(search_frame,command=self.fetch_data,text="Show All",width=8,font=("verdana",12,"bold"),fg="white",bg="navyblue")
        showAll_btn.grid(row=0,column=4,padx=5,pady=10,sticky=W)

        #Setting the Table Frame 
       
        #Searching System in Right Label Frame 
        table_frame = Frame(right_frame,bd=2,bg="white",relief=RIDGE)
        table_frame.place(x=10,y=90,width=635,height=360)

        #scroll bar 
        scroll_x = ttk.Scrollbar(table_frame,orient=HORIZONTAL)
        scroll_y = ttk.Scrollbar(table_frame,orient=VERTICAL)

        #creating the table 
        self.student_table = ttk.Treeview(table_frame,column=("Name","Roll-No","ID","Course","Year","Sem","Div","Gender","DOB","Mob-No","Address","Dep","Email","Teacher","Photo"),xscrollcommand=scroll_x.set,yscrollcommand=scroll_y.set)

        scroll_x.pack(side=BOTTOM,fill=X)
        scroll_y.pack(side=RIGHT,fill=Y)
        scroll_x.config(command=self.student_table.xview)
        scroll_y.config(command=self.student_table.yview)

        self.student_table.heading("Name",text="Name")
        self.student_table.heading("Roll-No",text="Roll-No")
        self.student_table.heading("ID",text="StudentID")
        self.student_table.heading("Course",text="Course")
        self.student_table.heading("Year",text="Year")
        self.student_table.heading("Sem",text="Semester")
        self.student_table.heading("Div",text="Division")
        self.student_table.heading("Gender",text="Gender")
        self.student_table.heading("DOB",text="DOB")
        self.student_table.heading("Mob-No",text="Mob-No")
        self.student_table.heading("Address",text="Address")
        self.student_table.heading("Dep",text="Department")
        self.student_table.heading("Email",text="Email")
        self.student_table.heading("Teacher",text="Teacher_Name")
        self.student_table.heading("Photo",text="PhotoSample")
        self.student_table["show"]="headings"


        # Setting Width of the Columns  
        self.student_table.column("Name",width=100)
        self.student_table.column("Roll-No",width=100)
        self.student_table.column("ID",width=100)
        self.student_table.column("Course",width=100)
        self.student_table.column("Year",width=100)
        self.student_table.column("Sem",width=100)
        self.student_table.column("Div",width=100)
        self.student_table.column("Gender",width=100)
        self.student_table.column("DOB",width=100)
        self.student_table.column("Mob-No",width=100)
        self.student_table.column("Address",width=100)
        self.student_table.column("Dep",width=100)
        self.student_table.column("Email",width=100)
        self.student_table.column("Teacher",width=100)
        self.student_table.column("Photo",width=100)


        self.student_table.pack(fill=BOTH,expand=1)
        self.student_table.bind("<ButtonRelease>",self.get_cursor)
        self.fetch_data()

    #Function Decleration
    def add_data(self):
        if self.var_dep.get()=="Select Department" or self.var_course.get=="Select Course" or self.var_year.get()=="Select Year" or self.var_semester.get()=="Select Semester" or self.var_std_id.get()=="" or self.var_std_name.get()=="" or self.var_div.get()=="" or self.var_roll.get()=="" or self.var_gender.get()=="" or self.var_dob.get()=="" or self.var_email.get()=="" or self.var_mob.get()=="" or self.var_address.get()=="" or self.var_teacher.get()=="":
            messagebox.showerror("Error","Please Fill All Fields are Required!",parent=self.root)
        else:
            try:
                conn = mysql.connector.connect(username='root', password='host',host='localhost',database='face_recognition',port=3306)
                mycursor = conn.cursor()
                mycursor.execute("insert into student values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",(
                self.var_std_id.get(),
                self.var_std_name.get(),
                self.var_roll.get(),
                self.var_course.get(),
                self.var_year.get(),
                self.var_semester.get(),
                self.var_div.get(),
                self.var_gender.get(),
                self.var_dob.get(),
                self.var_mob.get(),
                self.var_address.get(),
                self.var_dep.get(),
                self.var_email.get(),
                self.var_teacher.get(),
                self.var_radio1.get()
                ))

                conn.commit()
                self.fetch_data()
                conn.close()
                messagebox.showinfo("Success","All Records are Saved!",parent=self.root)
            except Exception as es:
                messagebox.showerror("Error",f"Due to: {str(es)}",parent=self.root)

    # Fetching data form database to table 

    def fetch_data(self):
        conn = mysql.connector.connect(username='root', password='host',host='localhost',database='face_recognition',port=3306)
        mycursor = conn.cursor() #Here you have to add your database credentials so that it should fetch data from the database

        mycursor.execute("select * from student")
        data=mycursor.fetchall()

        if len(data)!= 0:
            self.student_table.delete(*self.student_table.get_children())
            for i in data:
                self.student_table.insert("",END,values=i)
            conn.commit()
        conn.close()

    #Get cursor function

    def get_cursor(self,event=""):
        cursor_focus = self.student_table.focus()
        content = self.student_table.item(cursor_focus)
        data = content["values"]

        self.var_std_id.set(data[0]),
        self.var_std_name.set(data[1]),
        self.var_roll.set(data[2]),
        self.var_course.set(data[3]),
        self.var_year.set(data[4]),
        self.var_semester.set(data[5]),
        self.var_div.set(data[6]),
        self.var_gender.set(data[7]),
        self.var_dob.set(data[8]),
        self.var_mob.set(data[9]),
        self.var_address.set(data[10]),
        self.var_dep.set(data[11]),
        self.var_email.set(data[12]),
        self.var_teacher.set(data[13]),
        self.var_radio1.set(data[14])
    
    #Update Function
    def update_data(self):
        if self.var_dep.get()=="Select Department" or self.var_course.get=="Select Course" or self.var_year.get()=="Select Year" or self.var_semester.get()=="Select Semester" or self.var_std_id.get()=="" or self.var_std_name.get()=="" or self.var_div.get()=="" or self.var_roll.get()=="" or self.var_gender.get()=="" or self.var_dob.get()=="" or self.var_email.get()=="" or self.var_mob.get()=="" or self.var_address.get()=="" or self.var_teacher.get()=="":
            messagebox.showerror("Error","Please Fill All Fields are Required!",parent=self.root)
        else:
            try:
                Update=messagebox.askyesno("Update","Do you want to Update this Student Details!",parent=self.root)
                if Update > 0:
                    conn = mysql.connector.connect(username='root', password='host',host='localhost',database='face_recognition',port=3306)
                    mycursor = conn.cursor()
                    mycursor.execute("update student set Name=%s,Department=%s,Course=%s,Year=%s,Semester=%s,Division=%s,Gender=%s,DOB=%s,Mobile_No=%s,Address=%s,Roll_No=%s,Email=%s,Teacher_Name=%s,PhotoSample=%s where Student_ID=%s",( 
                    self.var_std_name.get(),
                    self.var_dep.get(),
                    self.var_course.get(),
                    self.var_year.get(),
                    self.var_semester.get(),
                    self.var_div.get(),
                    self.var_gender.get(),
                    self.var_dob.get(),
                    self.var_mob.get(),
                    self.var_address.get(),
                    self.var_roll.get(),
                    self.var_email.get(),
                    self.var_teacher.get(),
                    self.var_radio1.get(),
                    self.var_std_id.get()   
                    ))
                else:
                    if not Update:
                        return
                messagebox.showinfo("Success","Successfully Updated!",parent=self.root)
                conn.commit()
                self.fetch_data()
                conn.close()
            except Exception as es:
                messagebox.showerror("Error",f"Due to: {str(es)}",parent=self.root)
    
    #Delete Function
    def delete_data(self):
        if self.var_std_id.get()=="":
            messagebox.showerror("Error","Student Id Must be Required!",parent=self.root)
        else:
            try:
                delete=messagebox.askyesno("Delete","Do you want to Delete?",parent=self.root)
                if delete>0:
                    conn = mysql.connector.connect(username='root', password='host',host='localhost',database='face_recognition',port=3306)
                    mycursor = conn.cursor() 
                    sql="delete from student where Student_ID=%s"
                    val=(self.var_std_id.get(),)
                    mycursor.execute(sql,val)
                else:
                    if not delete:
                        return

                conn.commit()
                self.fetch_data()
                conn.close()
                messagebox.showinfo("Delete","Successfully Deleted!",parent=self.root)
            except Exception as es:
                messagebox.showerror("Error",f"Due to: {str(es)}",parent=self.root)    

    # Reset Function 
    def reset_data(self):
        self.var_std_id.set(""),
        self.var_std_name.set(""),
        self.var_roll.set(""),
        self.var_course.set("Select Course"),
        self.var_year.set("Select Year"),
        self.var_semester.set("Select Semester"),
        self.var_div.set("Morning"),
        self.var_gender.set("Male"),
        self.var_dob.set(""),
        self.var_mob.set(""),
        self.var_address.set(""),
        self.var_dep.set("Select Department"),
        self.var_email.set(""),
        self.var_teacher.set(""),
        self.var_radio1.set("")
    
    # Search Data
    def search_data(self):
        if self.var_search.get()=="" or self.var_searchTX.get()=="Select":
            messagebox.showerror("Error","Select Combo option and enter entry box",parent=self.root)
        else:
            try:
                conn = mysql.connector.connect(username='root', password='host',host='localhost',database='face_recognition',port=3306)
                my_cursor = conn.cursor()
                sql = "SELECT Student_ID,Name,Roll_No,Course,Year,Semester,Division,Gender,DOB,Mobile_No,Address,Department,Email,Teacher_Name,PhotoSample FROM student where Roll_No='" +str(self.var_search.get()) + "'" 
                my_cursor.execute(sql)
                # my_cursor.execute("select * from student where Roll_No= " +str(self.var_search.get())+" "+str(self.var_searchTX.get())+"")
                rows=my_cursor.fetchall()        
                if len(rows)!=0:
                    self.student_table.delete(*self.student_table.get_children())
                    for i in rows:
                        self.student_table.insert("",END,values=i)
                    if rows==None:
                        messagebox.showerror("Error","Data Not Found",parent=self.root)
                        conn.commit()
                conn.close()
            except Exception as es:
                messagebox.showerror("Error",f"Due To :{str(es)}",parent=self.root)


#This is use for upload the data to the database 

    def generate_dataset(self):
        if self.var_dep.get()=="Select Department" or self.var_course.get=="Select Course" or self.var_year.get()=="Select Year" or self.var_semester.get()=="Select Semester" or self.var_std_id.get()=="" or self.var_std_name.get()=="" or self.var_div.get()=="" or self.var_roll.get()=="" or self.var_gender.get()=="" or self.var_dob.get()=="" or self.var_email.get()=="" or self.var_mob.get()=="" or self.var_address.get()=="" or self.var_teacher.get()=="":
            messagebox.showerror("Error","Please Fill All Fields are Required!",parent=self.root)
        else:
            try:
                
                conn = mysql.connector.connect(username='root', password='host',host='localhost',database='face_recognition',port=3306)
                mycursor = conn.cursor()
                mycursor.execute("select * from student")
                myreslut = mycursor.fetchall()
                id=0
                for x in myreslut:
                    id+=1

                mycursor.execute("update student set Name=%s,Department=%s,Course=%s,Year=%s,Semester=%s,Division=%s,Gender=%s,DOB=%s,Mobile_No=%s,Address=%s,Roll_No=%s,Email=%s,Teacher_Name=%s,PhotoSample=%s where Student_ID=%s",( 
                    self.var_std_name.get(),
                    self.var_dep.get(),
                    self.var_course.get(),
                    self.var_year.get(),
                    self.var_semester.get(),
                    self.var_div.get(),
                    self.var_gender.get(),
                    self.var_dob.get(),
                    self.var_mob.get(),
                    self.var_address.get(),
                    self.var_roll.get(),
                    self.var_email.get(),
                    self.var_teacher.get(),
                    self.var_radio1.get(),
                    self.var_std_id.get()==id+1   
                    ))
                conn.commit()
                self.fetch_data()
                self.reset_data()
                conn.close()

                face_classifier = cv2.CascadeClassifier("d:/Python-FYP-Face-Recognition-Attendence-System-master/haarcascade_frontalface_default.xml")

                def face_croped(img):
                    # conver gary sacle
                    gray = cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
                    faces = face_classifier.detectMultiScale(gray,1.3,5)
                    #Scaling factor 1.3
                    # Minimum naber 5
                    for (x,y,w,h) in faces:
                        face_croped=img[y:y+h,x:x+w]
                        return face_croped
                cap=cv2.VideoCapture(0)
                img_id=0
                while True:
                    ret,my_frame=cap.read()
                    if face_croped(my_frame) is not None:
                        img_id+=1
                        face=cv2.resize(face_croped(my_frame),(200,200))
                        face=cv2.cvtColor(face,cv2.COLOR_BGR2GRAY)
                        file_path="D:/Python-FYP-Face-Recognition-Attendence-System-master/data_img/"+str(id)+"."+str(img_id)+".jpg"
                        cv2.imwrite(file_path,face)
                        cv2.putText(face,str(img_id),(50,50),cv2.FONT_HERSHEY_COMPLEX,2,(0,255,0),2)        
                        cv2.imshow("Capture Images",face)

                    if cv2.waitKey(1)==13 or int(img_id)==100:
                        break
                cap.release()
                cv2.destroyAllWindows()
                messagebox.showinfo("Result","Generating dataset completed!",parent=self.root)
            except Exception as es:
                messagebox.showerror("Error",f"Due to: {str(es)}",parent=self.root) 


if __name__ == "__main__":
    root=Tk()
    obj=Student(root)
    root.mainloop()


#TRAIN CODE
#This code is use to open and check how many faces we have and whose faces are present in our database

from sys import path
from tkinter import*
from tkinter import ttk
from PIL import Image,ImageTk
import os
import mysql.connector
import cv2
import numpy as np
from tkinter import messagebox
#This are the libraries and other files require to run the code properly.

class Train:

    def __init__(self, root):
        self.root = root
        self.root.geometry("1366x768+0+0")
        self.root.title("Train Panel")
        
        #In this section we are addinng the header,background,boxes and labels

        # Header image
        img = Image.open("d:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/mitname.jpg")  
        img = img.resize((1366, 130), Image.ANTIALIAS)
        self.photoimg = ImageTk.PhotoImage(img)

        # Setting image as a label
        f_lb1 = Label(self.root, image=self.photoimg)
        f_lb1.place(x=0, y=0, width=1366, height=130)

        # Background image
        bg1 = Image.open("D:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/t_bg1.jpg") 
        bg1 = bg1.resize((1366, 768), Image.ANTIALIAS)
        self.photobg1 = ImageTk.PhotoImage(bg1)

        # Setting image as a label
        bg_img = Label(self.root, image=self.photobg1)
        bg_img.place(x=0, y=130, width=1366, height=768)

        # Title section
        title_lb1 = Label(bg_img, text="Welcome to Training Panel", font=("verdana", 30, "bold"), bg="white", fg="navyblue")
        title_lb1.place(x=0, y=0, width=1366, height=45)

        # Creating buttons in below section
        # Training button 
        std_img_btn = Image.open("D:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/t_btn1.png") 
        std_img_btn = std_img_btn.resize((180, 180), Image.ANTIALIAS)
        self.std_img1 = ImageTk.PhotoImage(std_img_btn)

        std_b1 = Button(bg_img, image=self.std_img1, cursor="hand2")
        std_b1.place(x=600, y=170, width=180, height=180)

        std_b1_1 = Button(bg_img, text="Train Dataset", cursor="hand2", font=("tahoma", 15, "bold"),
                          bg="white", fg="navyblue", command=self.train_dataset)
        std_b1_1.place(x=600, y=350, width=180, height=45)

    def train_dataset(self):
        os.startfile('D:/studentsdetails/faces') #enter the dataset location here


if __name__ == "__main__":
    root = Tk()
    obj = Train(root)
    root.mainloop()
    
#HELP AND SUPPORT CODE

from tkinter import*
from PIL import Image,ImageTk
import webbrowser
#This are the libraries and other files require to run the code properly.

class Helpsupport:
    def __init__(self,root):
        self.root=root
        self.root.geometry("1366x768+0+0")
        self.root.title("Face_Recogonition_System")
        
        #Setting up the Image labels, Buttons and adding background image 

        # Header image  
        img=Image.open(r"d:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/mitname.jpg")
        img=img.resize((1366,130),Image.ANTIALIAS)
        self.photoimg=ImageTk.PhotoImage(img)

        # Setting image as a lable
        f_lb1 = Label(self.root,image=self.photoimg)
        f_lb1.place(x=0,y=0,width=1366,height=130)

        # Backgorund image 
        bg1=Image.open(r"d:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/bg4.png")
        bg1=bg1.resize((1366,768),Image.ANTIALIAS)
        self.photobg1=ImageTk.PhotoImage(bg1)

        # Setting image as a lable
        bg_img = Label(self.root,image=self.photobg1)
        bg_img.place(x=0,y=130,width=1366,height=768)


        #title section
        title_lb1 = Label(bg_img,text="Help Support",font=("verdana",30,"bold"),bg="white",fg="navyblue")
        title_lb1.place(x=0,y=0,width=1366,height=45)

    
        # Student button
        std_img_btn=Image.open(r"D:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/web.png")
        std_img_btn=std_img_btn.resize((180,180),Image.ANTIALIAS)
        self.std_img1=ImageTk.PhotoImage(std_img_btn)

        std_b1 = Button(bg_img,command=self.website,image=self.std_img1,cursor="hand2")
        std_b1.place(x=250,y=200,width=180,height=180)

        std_b1_1 = Button(bg_img,command=self.website,text="Website",cursor="hand2",font=("tahoma",15,"bold"),bg="white",fg="navyblue")
        std_b1_1.place(x=250,y=380,width=180,height=45)

        # Detect Face button 
        det_img_btn=Image.open(r"D:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/fb.png")
        det_img_btn=det_img_btn.resize((180,180),Image.ANTIALIAS)
        self.det_img1=ImageTk.PhotoImage(det_img_btn)

        det_b1 = Button(bg_img,command=self.facebook,image=self.det_img1,cursor="hand2",)
        det_b1.place(x=480,y=200,width=180,height=180)

        det_b1_1 = Button(bg_img,command=self.facebook,text="Facebook",cursor="hand2",font=("tahoma",15,"bold"),bg="white",fg="navyblue")
        det_b1_1.place(x=480,y=380,width=180,height=45)

         # Attendance System button 
        att_img_btn=Image.open(r"D:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/yt.png")
        att_img_btn=att_img_btn.resize((180,180),Image.ANTIALIAS)
        self.att_img1=ImageTk.PhotoImage(att_img_btn)

        att_b1 = Button(bg_img,command=self.youtube,image=self.att_img1,cursor="hand2",)
        att_b1.place(x=710,y=200,width=180,height=180)

        att_b1_1 = Button(bg_img,command=self.youtube,text="Youtube",cursor="hand2",font=("tahoma",15,"bold"),bg="white",fg="navyblue")
        att_b1_1.place(x=710,y=380,width=180,height=45)

         # Help Support button 
        hlp_img_btn=Image.open(r"D:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/gmail.png")
        hlp_img_btn=hlp_img_btn.resize((180,180),Image.ANTIALIAS)
        self.hlp_img1=ImageTk.PhotoImage(hlp_img_btn)

        hlp_b1 = Button(bg_img,command=self.gmail,image=self.hlp_img1,cursor="hand2",)
        hlp_b1.place(x=940,y=200,width=180,height=180)

        hlp_b1_1 = Button(bg_img,command=self.gmail,text="Gmail",cursor="hand2",font=("tahoma",15,"bold"),bg="white",fg="navyblue")
        hlp_b1_1.place(x=940,y=380,width=180,height=45)
  
  #Previous code of buttons 

        
        
    # Now creating function for button 
    def website(self):
        self.new = 1
        self.url = "https://exceleprep.com/"
        webbrowser.open(self.url,new=self.new)
    
    def facebook(self):
        self.new = 1
        self.url = "https://www.facebook.com/"
        webbrowser.open(self.url,new=self.new)
    
    def youtube(self):
        self.new = 1
        self.url = "https://www.youtube.com/channel/UCwpFCX_Z4SVkAT_6hPeUnsA"
        webbrowser.open(self.url,new=self.new)
    
    def gmail(self):
        self.new = 1
        self.url = "https://www.gmail.com"
        webbrowser.open(self.url,new=self.new)



if __name__ == "__main__":
    root=Tk()
    obj=Helpsupport(root)
    root.mainloop()
    
#DEVELOPER CODE

from tkinter import*
from tkinter import ttk
from train import Train
from PIL import Image,ImageTk
from student import Student
from train import Train
from face_recognition import Face_Recognition
from attendance import Attendance
import os
#This are the libraries and other files require to run the code properly.

class Developer:
    def __init__(self,root):
        self.root=root
        self.root.geometry("1366x768+0+0")
        self.root.title("Face_Recogonition_System")

#Setting up the Image labels, Buttons and adding background image
        
        # Header image  
        img=Image.open(r"d:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/mitname.jpg")
        img=img.resize((1366,130),Image.ANTIALIAS)
        self.photoimg=ImageTk.PhotoImage(img)

        # Setting image as a lable
        f_lb1 = Label(self.root,image=self.photoimg)
        f_lb1.place(x=0,y=0,width=1366,height=130)

        # Backgorund image 
        bg1=Image.open(r"D:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/bg4.png")
        bg1=bg1.resize((1366,768),Image.ANTIALIAS)
        self.photobg1=ImageTk.PhotoImage(bg1)

        # Setting image as a lable
        bg_img = Label(self.root,image=self.photobg1)
        bg_img.place(x=0,y=130,width=1366,height=768)


        #title section
        title_lb1 = Label(bg_img,text="Developer Pannel",font=("verdana",30,"bold"),bg="white",fg="navyblue")
        title_lb1.place(x=0,y=0,width=1366,height=45)

#Creating Buttons
        
        # Student button 
        std_img_btn=Image.open(r"d:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/eren.jpg")
        std_img_btn=std_img_btn.resize((180,180),Image.ANTIALIAS)
        self.std_img1=ImageTk.PhotoImage(std_img_btn)

        std_b1 = Button(bg_img,image=self.std_img1,cursor="hand2")
        std_b1.place(x=250,y=200,width=180,height=180)

        std_b1_1 = Button(bg_img,text="Eren eygaer",cursor="hand2",font=("tahoma",15,"bold"),bg="white",fg="navyblue")
        std_b1_1.place(x=250,y=380,width=180,height=45)

        # Button for face detection
        det_img_btn=Image.open(r"d:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/naruto.jpg")
        det_img_btn=det_img_btn.resize((180,180),Image.ANTIALIAS)
        self.det_img1=ImageTk.PhotoImage(det_img_btn)

        det_b1 = Button(bg_img,image=self.det_img1,cursor="hand2",)
        det_b1.place(x=480,y=200,width=180,height=180)

        det_b1_1 = Button(bg_img,text="Uzamaki Naruto",cursor="hand2",font=("tahoma",15,"bold"),bg="white",fg="navyblue")
        det_b1_1.place(x=480,y=380,width=180,height=45)

        # Button for Attendance System 
        att_img_btn=Image.open(r"d:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/tanjiro.jpg")
        att_img_btn=att_img_btn.resize((180,180),Image.ANTIALIAS)
        self.att_img1=ImageTk.PhotoImage(att_img_btn)

        att_b1 = Button(bg_img,image=self.att_img1,cursor="hand2",)
        att_b1.place(x=710,y=200,width=180,height=180)

        att_b1_1 = Button(bg_img,text="Tanjiro",cursor="hand2",font=("tahoma",15,"bold"),bg="white",fg="navyblue")
        att_b1_1.place(x=710,y=380,width=180,height=45)

        # Help Support button 
        hlp_img_btn=Image.open(r"d:/Python-FYP-Face-Recognition-Attendence-System-master/Images_GUI/zenistu.jpg")
        hlp_img_btn=hlp_img_btn.resize((180,180),Image.ANTIALIAS)
        self.hlp_img1=ImageTk.PhotoImage(hlp_img_btn)

        hlp_b1 = Button(bg_img,image=self.hlp_img1,cursor="hand2",)
        hlp_b1.place(x=940,y=200,width=180,height=180)

        hlp_b1_1 = Button(bg_img,text="Rengoku",cursor="hand2",font=("tahoma",15,"bold"),bg="white",fg="navyblue")
        hlp_b1_1.place(x=940,y=380,width=180,height=45)




if __name__ == "__main__":
    root=Tk()
    obj=Developer(root)
    root.mainloop()

#FACE RECOGNITION CODE
#This code recognize the face and mark there identity on excel sheet

import os
import cv2
import face_recognition
import numpy as np
import openpyxl
import pymysql
import datetime
#This are the libraries and other files require to run the code properly.

#In this section we are setting the database image means from where it should take the sample image and what should be its format
def get_encoded_faces():
    encoded = {}
    for dirpath, dnames, fnames in os.walk("./faces"):
        for f in fnames:
            if f.endswith(".jpg") or f.endswith(".png"):
                face = face_recognition.load_image_file("faces/" + f)
                encoding = face_recognition.face_encodings(face)[0]
                encoded[f.split(".")[0]] = encoding
    return encoded

def unknown_image_encoded(img_path):
    face = face_recognition.load_image_file("D:/studentsdetails/faces/" + img_path) #Added the path of the database folder where photo of everyone is available which can be use as sample image to compare
    encoding = face_recognition.face_encodings(face)[0]
    return encoding

def classify_face(im):
    faces = get_encoded_faces()
    faces_encoded = list(faces.values())
    known_face_names = list(faces.keys())

    img = cv2.imread(im, 1)

    face_locations = face_recognition.face_locations(img)
    unknown_face_encodings = face_recognition.face_encodings(img, face_locations)

    if len(unknown_face_encodings) == 0:
        print("No faces detected in the unknown image.")
        return []

    face_names = []
    for face_encoding in unknown_face_encodings:
        matches = face_recognition.compare_faces(faces_encoded, face_encoding)
        name = "unknown"   #If the face does not match with the database image it will display unknown on the face

        face_distances = face_recognition.face_distance(faces_encoded, face_encoding)
        best_match_index = np.argmin(face_distances)
        if matches[best_match_index]:
            name = known_face_names[best_match_index]

        face_names.append(name)

        for (top, right, bottom, left), name in zip(face_locations, face_names):
            cv2.rectangle(img, (left-20, top-20), (right+20, bottom+20), (255, 0, 0), 2)
            cv2.rectangle(img, (left-20, bottom-15), (right+20, bottom+20), (255, 0, 0), cv2.FILLED)
            font = cv2.FONT_HERSHEY_DUPLEX
            cv2.putText(img, name, (left - 20, bottom + 15), font, 1.0, (255, 255, 255), 2)

    #Show the image with recognized names
    cv2.imshow('Video', img)
    cv2.waitKey(0)  
    cv2.destroyAllWindows() 

    # Write the recognized names and every thing like id, roll-no into an Excel sheet
    write_to_excel(im, face_names)

    return face_names

def get_student_info(name):
    connection = pymysql.connect(
        host='localhost',  # Replace with the actual database host
        user='root',  # Replace with your database username
        password='host',  # Replace with your database password
        db='face_recognition',  # Replace with your database name
        charset='utf8mb4',
        cursorclass=pymysql.cursors.DictCursor
    )

    try:
        with connection.cursor() as cursor:
            sql = "SELECT rollno, id FROM student WHERE name = %s"
            cursor.execute(sql, (name,))  #It will select roll-no, id from student according to the name
            result = cursor.fetchone()
            return result
    finally:
        connection.close()



def write_to_excel(img_path, face_names):

#Add the excel sheet file here 
    excel_file = "d:/studentsdetails/recognized_faces.xlsx"
    if os.path.isfile(excel_file):
        workbook = openpyxl.load_workbook(excel_file)
        worksheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.append([ "Recognized Name", "Roll Number", "ID", "Date", "Time", "Attendance"]) #According to this format it will add everything in the Excel Sheet

    # Get the filename from the image path
    img_filename = os.path.basename(img_path)

    # Get the current date and time
    current_date = datetime.datetime.now().strftime('%Y-%m-%d')
    current_time = datetime.datetime.now().strftime('%H:%M:%S')

    # Keep track of recognized names that have been written. It will not add the same again once it is written in excel sheet
    written_names = set()

    # Write the recognized face names into the Excel sheet along with roll number and ID
    for name in face_names:
        if name != "unknown" and name not in written_names:
            student_info = get_student_info(name)
            if student_info:
                rollno = student_info['rollno']
                student_id = student_info['id']
                # Exclude the image filename from being written to the Excel sheet
                worksheet.append([name, rollno, student_id, current_date, current_time, "Present"])
                written_names.add(name)  # Mark the name as written

    # Save the changes to the Excel file
    workbook.save(excel_file)

face_names = classify_face("d:/studentsdetails/test.jpg") #Added the path of the image in which on which we want to perform the face recognition

#ATTENDANCE MARKING CODE
#This code forward all the data of excel sheet to the database 

import os
import mysql.connector
import openpyxl
from tkinter import *
from tkinter import filedialog, messagebox
#This are the libraries and other files require to run the code properly.

#Global variable for importing the data
mydata = []

class Attendance:

    def __init__(self, root):
        self.root = root
        self.root.geometry("600x400")
        self.root.title("Import Data to Database")

        import_btn = Button(self.root, text="Import XLSX", command=self.import_data)
        import_btn.pack(pady=20)

    def fetch_data(self):
        try:
            if len(mydata) < 1:
                messagebox.showerror("Error", "No Data Found!")
                return

            conn = mysql.connector.connect(user='root', password='host', host='localhost', database='face_recognition', port=3306)
            mycursor = conn.cursor()

            for row in mydata:
                mycursor.execute("INSERT INTO stdattendance (Name, std_roll_no, std_id, Date, Time, std_attendance) VALUES (%s, %s, %s, %s, %s, %s)",
                                 (row[0], row[1], row[2], row[3], row[4], row[5]))

            conn.commit()
            conn.close()
            messagebox.showinfo("Success", "Data Successfully Imported to Database!")

        except Exception as es:
            messagebox.showerror("Error", f"An error occurred: {str(es)}")

    def import_data(self):
        try:
            mydata.clear()
            fln = filedialog.askopenfilename(initialdir=os.getcwd(), title="Open XLSX",
                                             filetypes=(("XLSX File", "*.xlsx"), ("All File", "*.*")))
            workbook = openpyxl.load_workbook(fln)
            worksheet = workbook.active

            for row in worksheet.iter_rows(values_only=True):
                mydata.append(row)

            self.fetch_data()

        except Exception as es:
            messagebox.showerror("Error", f"An error occurred: {str(es)}")

if __name__ == "__main__":
    root = Tk()
    obj = Attendance(root)
    root.mainloop()
